# -*- coding: utf-8 -*-
"""Reconciliación FILA-A-FILA del motor del semáforo contra el Excel oficial.

Lee la hoja AVISOS del Excel de semáforos del equipo SAC, mapea sus columnas
crudas a los nombres canónicos de la app, corre sem_engine.compute_alerts y
compara, para cada aviso y cada una de las 7 etapas, el semáforo calculado por
la app contra el semáforo cacheado por el Excel (columnas BL..BX).

Uso (en LENOVO):
    python tools/reconciliar_semaforo.py "C:/ruta/Dashboard_..._SEMAFOROS.xlsx"

Sale con código 0 si TODO coincide (0 discrepancias), 1 si hay alguna.
La fecha de corte se toma de la celda BY2 del propio Excel (FECHA DE REPORTE).
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import pandas as pd
import numpy as np
from sem_engine import compute_alerts

# Excel(letra de columna) -> nombre canónico de la app
COL_MAP = {
    "P":  "FECHA_AVISO",
    "S":  "FECHA_ATENCION",
    "T":  "FECHA_PROGRAMACION_AJUSTE",
    "V":  "FECHA_AJUSTE_ACTA_1",
    "W":  "ESTADO_SINIESTRO",
    "Y":  "FECHA_REPROGRAMACION_01",
    "Z":  "FECHA_REPROGRAMACION_02",
    "AA": "FECHA_REPROGRAMACION_03",
    "AB": "FECHA_REPROGRAMACION_04",
    "AC": "FECHA_REPROGRAMACION_05",
    "AD": "FECHA_REPROGRAMACION_06",
    "AE": "FECHA_AJUSTE_ACTA_FINAL",
    "AF": "ESTADO_INSPECCION",
    "AO": "DICTAMEN",
    "AT": "DUPLIC RESULT",
    "AY": "CODIGO_PADRON",
    "AZ": "FECHA_ENVIO_DRAS",
    "BA": "FECHA_VALIDACION",
    "BB": "FECHA_DESEMBOLSO",
    "BF": "OBSERVACION",
}
# Etapa app (sufijo 01..07) -> columna de semáforo cacheado del Excel
SEM_CACHE_COL = {"01": "BL", "02": "BN", "03": "BP", "04": "BR",
                 "05": "BT", "06": "BV", "07": "BX"}
NOMBRE = {"01": "ATENCIÓN", "02": "PROGRAMACIÓN", "03": "AJUSTE 01",
          "04": "REPROGRAMACIÓN", "05": "PADRÓN", "06": "VALIDACIÓN",
          "07": "PAGO SAC"}


def _norm_sem(v):
    """Normaliza un semáforo a {-1,0,1,2,3} o None (blanco)."""
    if v is None:
        return None
    if isinstance(v, float) and np.isnan(v):
        return None
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
        try:
            v = float(v)
        except ValueError:
            return None
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None


def cargar_avisos(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["AVISOS"]
    maxr = ws.max_row
    fecha_reporte = ws["BY2"].value
    cols = {name: [ws[f"{letter}{r}"].value for r in range(2, maxr + 1)]
            for letter, name in COL_MAP.items()}
    df = pd.DataFrame(cols)
    sem_xl = {suf: [ws[f"{col}{r}"].value for r in range(2, maxr + 1)]
              for suf, col in SEM_CACHE_COL.items()}
    return df, sem_xl, fecha_reporte


def reconciliar(xlsx_path):
    df, sem_xl, fecha_reporte = cargar_avisos(xlsx_path)
    today = pd.Timestamp(fecha_reporte).normalize()
    print(f"Archivo : {os.path.basename(xlsx_path)}")
    print(f"Filas   : {len(df)}")
    print(f"Corte   : {today.date()} (celda BY2)\n")

    out = compute_alerts(df, today=today)

    total_mismatch = 0
    print(f"{'ETAPA':<18}{'OK':>8}{'DISCREP':>9}{'% MATCH':>9}")
    print("-" * 44)
    detalle = {}
    for suf in ["01", "02", "03", "04", "05", "06", "07"]:
        app_vals = [_norm_sem(v) for v in out[f"SEMAFORO_{suf}"]]
        xl_vals = [_norm_sem(v) for v in sem_xl[suf]]
        diffs = [(i, a, x) for i, (a, x) in enumerate(zip(app_vals, xl_vals)) if a != x]
        ok = len(df) - len(diffs)
        pct = 100.0 * ok / len(df) if len(df) else 100.0
        total_mismatch += len(diffs)
        detalle[suf] = diffs
        print(f"A{suf} {NOMBRE[suf]:<14}{ok:>8}{len(diffs):>9}{pct:>8.2f}%")

    print("-" * 44)
    print(f"{'TOTAL discrepancias':<27}{total_mismatch:>8}\n")

    # Mostrar hasta 10 ejemplos de discrepancia por etapa
    for suf, diffs in detalle.items():
        if not diffs:
            continue
        print(f"\n  Ejemplos A{suf} {NOMBRE[suf]} (fila Excel = i+2):")
        from collections import Counter
        combo = Counter((a, x) for _, a, x in diffs)
        for (a, x), n in combo.most_common(8):
            print(f"    app={a!s:>5}  excel={x!s:>5}  ->  {n} filas")
        for i, a, x in diffs[:5]:
            print(f"    fila {i+2}: app={a} excel={x}  "
                  f"alerta_app={out[f'ALERTA_{suf}'].iloc[i]!r}")

    return total_mismatch


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python tools/reconciliar_semaforo.py <ruta_xlsx>")
        sys.exit(2)
    n = reconciliar(sys.argv[1])
    print("\n" + ("✅ RECONCILIACIÓN PERFECTA (0 discrepancias)" if n == 0
                  else f"❌ {n} discrepancias — revisar arriba"))
    sys.exit(0 if n == 0 else 1)
