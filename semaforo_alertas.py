"""
Semáforo de Alertas SAC — Motor de cálculo y UI Streamlit.
Sistema de control de plazos para las 7 etapas del flujo del
Seguro Agrícola Catastrófico (SAC). El cálculo por etapa vive en
sem_engine.compute_alerts (port fiel del Excel oficial, reconciliado
fila-a-fila — ver tests/test_reconciliacion_semaforo.py).
"""

import io
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Cache decorator con fallback: si Streamlit no está en runtime (tests,
# CLI), @_cache_data se vuelve un no-op y el cálculo corre directo.
try:
    _cache_data = st.cache_data
except Exception:  # pragma: no cover
    def _cache_data(*dargs, **dkwargs):
        def _dec(fn):
            return fn
        if dargs and callable(dargs[0]):
            return dargs[0]
        return _dec

# ═══════════════════════════════════════════════════════════════
#  CONSTANTES
# ═══════════════════════════════════════════════════════════════

COLS_MINIMAS = ["FECHA_AVISO", "FECHA_ATENCION"]

STAGES = [
    {"key": "atencion",        "label": "Atención",        "icon": "1", "emoji": "📋"},
    {"key": "programacion",    "label": "Programación",    "icon": "2", "emoji": "📅"},
    {"key": "ajuste",          "label": "Ajuste 01",       "icon": "3", "emoji": "🔍"},
    {"key": "reprogramacion",  "label": "Reprogramación",  "icon": "4", "emoji": "🔄"},
    {"key": "padron",          "label": "Padrón",          "icon": "5", "emoji": "📝"},
    {"key": "validacion",      "label": "Validación",      "icon": "6", "emoji": "✅"},
    {"key": "pago",            "label": "Pago SAC",        "icon": "7", "emoji": "💰"},
]

# Etapa (key UI) → columna de semáforo del motor (SEMAFORO_01..07).
STAGE_SEM_COL = {
    "atencion": "SEMAFORO_01", "programacion": "SEMAFORO_02",
    "ajuste": "SEMAFORO_03", "reprogramacion": "SEMAFORO_04",
    "padron": "SEMAFORO_05", "validacion": "SEMAFORO_06", "pago": "SEMAFORO_07",
}

COLORS = {
    "verde":  {"hex": "#27ae60", "bg": "#d4edda", "label": "En plazo"},
    "ambar":  {"hex": "#f39c12", "bg": "#fff3cd", "label": "Riesgo"},
    "rojo":   {"hex": "#e74c3c", "bg": "#f8d7da", "label": "Vencido"},
}

# Mapeo precomputado (evita recrear dict en cada render)
_STAGE_KEY_TO_LABEL = {s["key"]: s["label"] for s in STAGES}


# ═══════════════════════════════════════════════════════════════
#  A) MOTOR DE CÁLCULO (vectorizado)
# ═══════════════════════════════════════════════════════════════

def check_semaforo_columns(df):
    """Verifica si el DataFrame tiene las columnas mínimas para el semáforo."""
    return all(c in df.columns for c in COLS_MINIMAS)


def _safe_col(df, col):
    """Retorna la columna como Series si existe, si no retorna NaT/NaN Series."""
    if col in df.columns:
        return df[col]
    return pd.Series(pd.NaT, index=df.index)


def compute_semaforo(df, today=None, cache_key=None):
    """
    Calcula las 7 etapas independientes (ALERTA 01..07 + SEMAFORO 01..07)
    según las reglas oficiales del equipo SAC, y deriva las columnas
    de resumen SEM_* (peor caso) para compatibilidad con la UI existente.

    Las 7 etapas se calculan en sem_engine.compute_alerts(). Esta función
    luego deriva:
      - SEM_ETAPA: la etapa con la peor alerta (más crítica) o
                   "completado" si no hay ninguna alerta activa.
      - SEM_ALERTA: el peor color entre las 6 alertas
                    (rojo > ámbar > verde > sin alerta).
      - SEM_DIAS: los días de la peor alerta (si aplica).
      - SEM_DETALLE: el texto descriptivo de la alerta crítica.

    Validado contra "Dashboard_SAC_25-26_..._SEMAFOROS.xlsx":
      - Reconciliación fila-a-fila: 7/7 etapas con 100% de match en el
        semáforo (-1/0/1/2/3) sobre las 12,914 filas reales del corte.
      - Días calendario en A01-A06 y días hábiles en A07, tal cual el Excel.

    PERFORMANCE: este es el cálculo más pesado de la app (7 etapas
    vectorizadas sobre ~13k filas) y corría en CADA render
    del dashboard (banner de alertas) y de la página del semáforo. Si se
    pasa `cache_key` (tupla pequeña y estable, p.ej. (fecha_corte,
    total_avisos)), el resultado se cachea por (cache_key, día). Sin
    cache_key corre directo (tests / llamadas sueltas).
    """
    if today is None:
        today = pd.Timestamp.now().normalize()
    if cache_key is None:
        return _compute_semaforo_impl(df, today)
    return _compute_semaforo_cached(df, cache_key, today)


@_cache_data(show_spinner=False, ttl=600)
def _compute_semaforo_cached(_df, cache_key, today):
    """Wrapper cacheado. `_df` lleva guion bajo → Streamlit NO lo hashea
    (sería caro sobre 11k filas). La identidad del cache la dan cache_key
    (cambia al recargar datos) y today (cambia cada día → cambian plazos)."""
    return _compute_semaforo_impl(_df, today)


def _compute_semaforo_impl(df, today):
    """Cálculo real de las 7 etapas + derivación de columnas SEM_* (peor caso)."""
    from sem_engine import compute_alerts

    # 1. Calcular las 7 etapas independientes
    result = compute_alerts(df, today=today)

    # 2. Derivar SEM_* desde las 7 etapas (peor caso)
    n = len(result)

    # Mapeo etapa → columnas de salida (7 etapas, A01..A07 del Excel)
    stages_info = [
        ("atencion",        "ALERTA_01",   "SEMAFORO_01"),
        ("programacion",    "ALERTA_02",   "SEMAFORO_02"),
        ("ajuste",          "ALERTA_03",   "SEMAFORO_03"),
        ("reprogramacion",  "ALERTA_04",   "SEMAFORO_04"),
        ("padron",          "ALERTA_05",   "SEMAFORO_05"),
        ("validacion",      "ALERTA_06",   "SEMAFORO_06"),
        ("pago",            "ALERTA_07",   "SEMAFORO_07"),
    ]

    sem_etapa   = pd.Series("completado", index=result.index, dtype="object")
    sem_alerta  = pd.Series("", index=result.index, dtype="object")
    sem_dias    = pd.Series(0, index=result.index, dtype="int64")
    sem_detalle = pd.Series("", index=result.index, dtype="object")
    worst_score = pd.Series(0, index=result.index, dtype="int64")
    # Score: 3 = rojo (worst), 2 = ámbar, 1 = verde, 0 = sin alerta

    color_map = {1.0: "verde", 2.0: "ambar", 3.0: "rojo"}

    for stage_key, alert_col, sem_col in stages_info:
        if alert_col not in result.columns:
            continue
        s = result[sem_col]
        s_int = pd.to_numeric(s, errors="coerce").fillna(0).astype(int)

        # Para cada fila: si su semáforo numérico es mayor que worst_score,
        # actualiza con esta etapa.
        is_worse = s_int > worst_score
        worst_score = worst_score.where(~is_worse, s_int)
        sem_etapa = sem_etapa.where(~is_worse, stage_key)
        sem_alerta = sem_alerta.where(~is_worse, s_int.map(color_map).fillna(""))
        # Días: extraer el número del string "(N días)" si existe
        text = result[alert_col].astype(str)
        days_extracted = text.str.extract(r"\((-?\d+)\s*d", expand=False)
        days_int = pd.to_numeric(days_extracted, errors="coerce").fillna(0).astype("int64")
        sem_dias = sem_dias.where(~is_worse, days_int)
        sem_detalle = sem_detalle.where(~is_worse, text)

    result["SEM_ETAPA"] = sem_etapa
    result["SEM_ALERTA"] = sem_alerta
    result["SEM_DIAS"] = sem_dias
    result["SEM_DETALLE"] = sem_detalle

    return result


def get_pipeline_summary(df_sem):
    """Resumen por etapa con conteo INDEPENDIENTE por etapa (igual que el
    Excel/R1): cada aviso cuenta en TODAS sus etapas, no solo en la peor.

    Por etapa devuelve verde/ámbar/rojo (alertas activas) + conforme/excluido
    (informativos). Esto es lo que debe coincidir con la hoja R1 del Excel.
    """
    summary = {}
    for s in STAGES:
        k = s["key"]
        col = STAGE_SEM_COL[k]
        if col not in df_sem.columns:
            summary[k] = {"verde": 0, "ambar": 0, "rojo": 0,
                          "conforme": 0, "excluido": 0, "total": 0}
            continue
        v = pd.to_numeric(df_sem[col], errors="coerce")
        verde = int((v == 1).sum()); ambar = int((v == 2).sum())
        rojo = int((v == 3).sum()); conforme = int((v == 0).sum())
        excluido = int((v == -1).sum())
        summary[k] = {
            "verde": verde, "ambar": ambar, "rojo": rojo,
            "conforme": conforme, "excluido": excluido,
            "total": verde + ambar + rojo,
        }
    return summary


def get_kpi_summary(df_sem):
    """KPIs globales."""
    active = df_sem[df_sem["SEM_ETAPA"] != "completado"]
    total = len(active)
    if total == 0:
        return {"total": 0, "verde": 0, "ambar": 0, "rojo": 0,
                "pct_verde": 0, "pct_ambar": 0, "pct_rojo": 0}
    v = int((active["SEM_ALERTA"] == "verde").sum())
    a = int((active["SEM_ALERTA"] == "ambar").sum())
    r = int((active["SEM_ALERTA"] == "rojo").sum())
    return {
        "total": total, "verde": v, "ambar": a, "rojo": r,
        "pct_verde": round(v / total * 100, 1),
        "pct_ambar": round(a / total * 100, 1),
        "pct_rojo":  round(r / total * 100, 1),
    }


def get_notification_banner_html(datos):
    """Genera HTML de banner de notificaciones con resumen de alertas.
    Retorna None si no hay columnas de semáforo."""
    df = datos.get("midagri")
    if df is None or df.empty:
        return None
    if not check_semaforo_columns(df):
        return None

    today = pd.Timestamp.now().normalize()
    _ck = (datos.get("fecha_corte", ""), datos.get("total_avisos", 0))
    df_sem = compute_semaforo(df, today, cache_key=_ck)
    kpis = get_kpi_summary(df_sem)

    if kpis["total"] == 0:
        return None

    return (
        '<div style="background:linear-gradient(135deg,#fff8f0,#fef3e6);border:1px solid #f0c78a;'
        'border-radius:12px;padding:0.8rem 1.2rem;margin-bottom:1rem;display:flex;'
        'align-items:center;gap:1.5rem;flex-wrap:wrap;">'
        '<div style="font-weight:700;color:#0c2340;font-size:0.85rem;">Alertas SAC</div>'
        f'<div style="display:flex;gap:0.3rem;align-items:center;">'
        f'<span style="background:#e74c3c;color:white;border-radius:20px;padding:2px 10px;'
        f'font-size:0.78rem;font-weight:700;">{kpis["rojo"]}</span>'
        f'<span style="color:#64748b;font-size:0.75rem;">vencidos</span></div>'
        f'<div style="display:flex;gap:0.3rem;align-items:center;">'
        f'<span style="background:#f39c12;color:white;border-radius:20px;padding:2px 10px;'
        f'font-size:0.78rem;font-weight:700;">{kpis["ambar"]}</span>'
        f'<span style="color:#64748b;font-size:0.75rem;">en riesgo</span></div>'
        f'<div style="display:flex;gap:0.3rem;align-items:center;">'
        f'<span style="background:#27ae60;color:white;border-radius:20px;padding:2px 10px;'
        f'font-size:0.78rem;font-weight:700;">{kpis["verde"]}</span>'
        f'<span style="color:#64748b;font-size:0.75rem;">en plazo</span></div>'
        '</div>'
    )


def generate_sankey_figure(df_sem):
    """Genera diagrama Sankey mostrando flujo de avisos por las 6 etapas."""
    active = df_sem[df_sem["SEM_ETAPA"] != "completado"]
    if active.empty:
        return None

    stage_keys = [s["key"] for s in STAGES]
    stage_labels = [s["label"] for s in STAGES] + ["Completado"]
    node_colors = ["#3498db", "#2980b9", "#1a5276", "#f39c12", "#e67e22",
                   "#16a085", "#27ae60", "#95a5a6"]

    # Contar avisos por etapa y alerta
    sources, targets, values, link_colors = [], [], [], []
    color_map = {"verde": "rgba(39,174,96,0.4)", "ambar": "rgba(243,156,18,0.4)", "rojo": "rgba(231,76,60,0.4)"}

    for i, key in enumerate(stage_keys):
        sub = active[active["SEM_ETAPA"] == key]
        for alert_level in ["verde", "ambar", "rojo"]:
            cnt = int((sub["SEM_ALERTA"] == alert_level).sum())
            if cnt > 0:
                # Flujo de esta etapa hacia "Completado" (último nodo)
                sources.append(i)
                targets.append(len(stage_keys))  # Completado
                values.append(cnt)
                link_colors.append(color_map.get(alert_level, "rgba(150,150,150,0.3)"))

    if not values:
        return None

    fig = go.Figure(go.Sankey(
        arrangement="snap",
        node=dict(
            pad=24, thickness=22,
            label=stage_labels,
            color=node_colors,
            line=dict(color="rgba(0,0,0,0)", width=0),
        ),
        link=dict(
            source=sources, target=targets, value=values, color=link_colors,
            hovertemplate="<b>%{source.label}</b> → <b>%{target.label}</b><br>Avisos: %{value:,}<extra></extra>",
        ),
    ))
    fig.update_layout(
        title=dict(
            text="<b>Flujo de Avisos por Etapas</b>"
                 "<br><span style='font-size:11px;color:#64748b;font-weight:400'>"
                 "Semáforo: 🟢 verde · 🟡 ámbar · 🔴 rojo</span>",
            font=dict(size=16, color="#0c2340", family="Segoe UI, Arial"),
            x=0.0, xanchor="left", y=0.97,
        ),
        font=dict(size=12, family="Segoe UI, Arial, sans-serif", color="#334155"),
        height=390,
        margin=dict(l=20, r=20, t=80, b=20),
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        hoverlabel=dict(bgcolor="#ffffff", bordercolor="#e2e8f0",
                        font=dict(size=12, family="Segoe UI", color="#0c2340")),
    )
    return fig


def export_semaforo_excel(df_sem, pipeline, kpis):
    """Genera Excel con formato condicional verde/ámbar/rojo."""
    wb = Workbook()

    # ── Hoja 1: Resumen Pipeline ──
    ws1 = wb.active
    ws1.title = "Resumen Pipeline"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2C3E50")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Etapa", "Verde", "Ámbar", "Rojo", "Total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    fill_v = PatternFill("solid", fgColor="27AE60")
    fill_a = PatternFill("solid", fgColor="F39C12")
    fill_r = PatternFill("solid", fgColor="E74C3C")

    for row_idx, s in enumerate(STAGES, 2):
        k = s["key"]
        data = pipeline.get(k, {"verde": 0, "ambar": 0, "rojo": 0, "total": 0})
        ws1.cell(row=row_idx, column=1, value=s["label"]).border = thin_border
        c_v = ws1.cell(row=row_idx, column=2, value=data["verde"])
        c_v.fill = fill_v if data["verde"] > 0 else PatternFill()
        c_v.font = Font(color="FFFFFF", bold=True) if data["verde"] > 0 else Font()
        c_v.border = thin_border
        c_v.alignment = Alignment(horizontal="center")
        c_a = ws1.cell(row=row_idx, column=3, value=data["ambar"])
        c_a.fill = fill_a if data["ambar"] > 0 else PatternFill()
        c_a.font = Font(color="FFFFFF", bold=True) if data["ambar"] > 0 else Font()
        c_a.border = thin_border
        c_a.alignment = Alignment(horizontal="center")
        c_r = ws1.cell(row=row_idx, column=4, value=data["rojo"])
        c_r.fill = fill_r if data["rojo"] > 0 else PatternFill()
        c_r.font = Font(color="FFFFFF", bold=True) if data["rojo"] > 0 else Font()
        c_r.border = thin_border
        c_r.alignment = Alignment(horizontal="center")
        ws1.cell(row=row_idx, column=5, value=data["total"]).border = thin_border

    # KPIs resumen
    row_k = len(STAGES) + 3
    ws1.cell(row=row_k, column=1, value="RESUMEN GLOBAL").font = Font(bold=True, size=12)
    ws1.cell(row=row_k+1, column=1, value="Total alertas activas")
    ws1.cell(row=row_k+1, column=2, value=kpis["total"])
    ws1.cell(row=row_k+2, column=1, value="% Verde")
    ws1.cell(row=row_k+2, column=2, value=f"{kpis['pct_verde']}%")
    ws1.cell(row=row_k+3, column=1, value="% Ámbar")
    ws1.cell(row=row_k+3, column=2, value=f"{kpis['pct_ambar']}%")
    ws1.cell(row=row_k+4, column=1, value="% Rojo")
    ws1.cell(row=row_k+4, column=2, value=f"{kpis['pct_rojo']}%")

    for c in range(1, 6):
        ws1.column_dimensions[chr(64 + c)].width = 18

    # ── Hoja 2: Detalle Alertas ──
    ws2 = wb.create_sheet("Detalle Alertas")

    display_cols = ["CODIGO_AVISO", "DEPARTAMENTO", "PROVINCIA", "DISTRITO",
                    "SECTOR_ESTADISTICO", "TIPO_CULTIVO", "EMPRESA",
                    "TIPO_SINIESTRO", "SEM_ETAPA", "SEM_ALERTA",
                    "SEM_DIAS", "SEM_DETALLE"]
    available = [c for c in display_cols if c in df_sem.columns]
    df_export = df_sem[df_sem["SEM_ETAPA"] != "completado"][available].copy()

    col_labels = {
        "CODIGO_AVISO": "Código Aviso", "DEPARTAMENTO": "Departamento",
        "PROVINCIA": "Provincia", "DISTRITO": "Distrito",
        "SECTOR_ESTADISTICO": "Sector", "TIPO_CULTIVO": "Cultivo",
        "EMPRESA": "Empresa",
        "TIPO_SINIESTRO": "Tipo Siniestro", "SEM_ETAPA": "Etapa",
        "SEM_ALERTA": "Alerta", "SEM_DIAS": "Días", "SEM_DETALLE": "Detalle"
    }

    for col_idx, col in enumerate(available, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_labels.get(col, col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    fill_map = {"verde": fill_v, "ambar": fill_a, "rojo": fill_r}
    font_white = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")

    # Batch write: usar .values en vez de iterrows (5-10x más rápido)
    alerta_col_idx = available.index("SEM_ALERTA") + 1 if "SEM_ALERTA" in available else -1
    data_matrix = df_export[available].values
    for row_idx, row_data in enumerate(data_matrix, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == alerta_col_idx and val in fill_map:
                cell.fill = fill_map[val]
                cell.font = font_white
                cell.alignment = center_align

    from openpyxl.utils import get_column_letter
    for i in range(1, len(available) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
#  B) UI STREAMLIT
# ═══════════════════════════════════════════════════════════════

def _css_semaforo():
    """CSS para el pipeline visual y badges."""
    return """
    <style>
    /* ── Info de fecha de corte ── */
    .sem-corte-info {
        display:flex; align-items:center; gap:10px; height:100%;
        background:#f4f8f5; border:1px solid #e1ebe4; border-radius:10px;
        padding:8px 14px; color:#3a4a41; font-size:12.5px;
    }
    .sem-corte-info .ms { font-size:22px; color:#408B14; }
    .sem-corte-info span { color:#7c8a82; font-size:11.5px; }
    /* ── Grid de etapas ── */
    .sem-grid {
        display:grid; grid-template-columns:repeat(auto-fit, minmax(168px, 1fr));
        gap:12px; margin:14px 0 4px;
    }
    .sem-card {
        background:#fff; border:1px solid #e6ece8; border-radius:14px;
        padding:13px 13px 9px; box-shadow:0 1px 3px rgba(16,40,28,.06);
        transition:transform .12s ease, box-shadow .12s ease;
    }
    .sem-card:hover { transform:translateY(-2px); box-shadow:0 6px 16px rgba(16,40,28,.10); }
    .sem-card-head { display:flex; align-items:center; gap:8px; margin-bottom:10px; }
    .sem-card-num {
        width:26px; height:26px; border-radius:50%; flex:0 0 auto;
        display:flex; align-items:center; justify-content:center;
        background:var(--accent); color:#fff; font-weight:800; font-size:13px;
        box-shadow:0 1px 3px rgba(0,0,0,.18);
    }
    .sem-card-name {
        font-size:12px; font-weight:700; color:#1f3d2b;
        text-transform:uppercase; letter-spacing:.3px; line-height:1.15;
    }
    .sem-bar {
        display:flex; height:8px; border-radius:6px; overflow:hidden;
        background:#eef2f0; margin-bottom:10px;
    }
    .sem-bar .sb { display:block; height:100%; }
    .sb-v { background:#27ae60; } .sb-a { background:#f39c12; } .sb-r { background:#e74c3c; }
    .sb-empty { background:#e2e8e5; }
    .sem-chips { display:flex; gap:6px; }
    .sem-chips .chip {
        flex:1; text-align:center; font-weight:800; font-size:14px;
        border-radius:8px; padding:5px 0; color:#fff; line-height:1.1;
    }
    .chip-v { background:#27ae60; } .chip-a { background:#f39c12; } .chip-r { background:#e74c3c; }
    .sem-card-foot {
        margin-top:8px; font-size:10.5px; color:#7c8a82; text-align:center;
    }
    .sem-kpi-row {
        display: flex; gap: 16px; margin: 16px 0;
        justify-content: center; flex-wrap: wrap;
    }
    .sem-kpi {
        background: #fff; border-radius: 10px; padding: 16px 24px;
        text-align: center; min-width: 140px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.07); border-left: 4px solid #408B14;
    }
    .sem-kpi-val { font-size: 28px; font-weight: 800; color: #2C3E50; }
    .sem-kpi-label { font-size: 11px; color: #7f8c8d; text-transform: uppercase; letter-spacing: 0.5px; }
    .sem-kpi.verde  { border-left-color: #27ae60; }
    .sem-kpi.ambar  { border-left-color: #f39c12; }
    .sem-kpi.rojo   { border-left-color: #e74c3c; }
    .sem-kpi.total  { border-left-color: #408B14; }
    .sem-drilldown-title {
        font-size: 13px; font-weight: 700; color: #2C3E50;
        margin: 8px 0 4px 0; padding: 6px 10px;
        background: #f8f9fa; border-radius: 6px; border-left: 3px solid #408B14;
    }
    </style>
    """


def _render_pipeline_html(pipeline):
    """Grid de 7 tarjetas por etapa: barra de proporción + chips V/Á/R +
    conformes. El acento del encabezado refleja el peor estado activo."""
    cards = []
    for s in STAGES:
        d = pipeline.get(s["key"], {"verde": 0, "ambar": 0, "rojo": 0,
                                    "conforme": 0, "excluido": 0, "total": 0})
        v, a, r = d["verde"], d["ambar"], d["rojo"]
        conf = d.get("conforme", 0)
        tot = v + a + r
        if tot > 0:
            pv, pa, pr = 100 * v / tot, 100 * a / tot, 100 * r / tot
            bar = (f'<span class="sb sb-v" style="width:{pv:.2f}%"></span>'
                   f'<span class="sb sb-a" style="width:{pa:.2f}%"></span>'
                   f'<span class="sb sb-r" style="width:{pr:.2f}%"></span>')
            foot = f'{tot:,} en proceso · {conf:,} conformes'
        else:
            bar = '<span class="sb sb-empty" style="width:100%"></span>'
            foot = f'sin alertas · {conf:,} conformes'
        accent = ("#e74c3c" if r > 0 else "#f39c12" if a > 0
                  else "#27ae60" if v > 0 else "#cbd5e1")
        cards.append(f'''
        <div class="sem-card">
          <div class="sem-card-head" style="--accent:{accent}">
            <span class="sem-card-num">{s["icon"]}</span>
            <span class="sem-card-name">{s["label"]}</span>
          </div>
          <div class="sem-bar">{bar}</div>
          <div class="sem-chips">
            <span class="chip chip-v" title="En plazo">{v:,}</span>
            <span class="chip chip-a" title="En riesgo">{a:,}</span>
            <span class="chip chip-r" title="Vencidos">{r:,}</span>
          </div>
          <div class="sem-card-foot">{foot}</div>
        </div>''')
    return '<div class="sem-grid">' + "".join(cards) + '</div>'


def _render_kpis_html(kpis):
    """Genera KPIs como HTML."""
    return f'''
    <div class="sem-kpi-row">
        <div class="sem-kpi total">
            <div class="sem-kpi-val">{kpis["total"]:,}</div>
            <div class="sem-kpi-label">Alertas Activas</div>
        </div>
        <div class="sem-kpi verde">
            <div class="sem-kpi-val">{kpis["pct_verde"]}%</div>
            <div class="sem-kpi-label">En Plazo ({kpis["verde"]:,})</div>
        </div>
        <div class="sem-kpi ambar">
            <div class="sem-kpi-val">{kpis["pct_ambar"]}%</div>
            <div class="sem-kpi-label">En Riesgo ({kpis["ambar"]:,})</div>
        </div>
        <div class="sem-kpi rojo">
            <div class="sem-kpi-val">{kpis["pct_rojo"]}%</div>
            <div class="sem-kpi-label">Vencidos ({kpis["rojo"]:,})</div>
        </div>
    </div>
    '''


def _stage_label(key):
    """Devuelve el label legible de una etapa."""
    for s in STAGES:
        if s["key"] == key:
            return f'{s["emoji"]} {s["label"]}'
    return key


def render_semaforo_tab(datos):
    """Punto de entrada: renderiza la pestaña Semáforo de Alertas."""
    st.markdown(_css_semaforo(), unsafe_allow_html=True)

    st.markdown("""
    <div style="background:#1f3d2b;
         padding:18px 24px;border-radius:10px;margin-bottom:18px;">
        <span style="color:#fff;font-size:22px;font-weight:700;">
        <span class="ms" style="color:inherit;">traffic</span> Semáforo de Alertas — Control de Plazos SAC</span><br>
        <span style="color:#d4edda;font-size:13px;">
        Monitoreo en tiempo real de las 7 etapas del Seguro Agrícola Catastrófico</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Explicación de criterios ──
    with st.expander("¿Cómo funciona el Semáforo? — Reglas por etapa", expanded=False):
        st.markdown("""
**El semáforo evalúa cada aviso en 7 etapas independientes del flujo SAC.** Cada etapa tiene su propio nivel de alerta (verde / ámbar / rojo) según el plazo transcurrido. El "Pipeline" cuenta cada etapa de forma independiente (igual que la hoja R1 del Excel); el indicador de "peor etapa" prioriza por aviso.

**Días calendario** en las etapas 1-6; **días hábiles** (sáb+dom y feriados no laborables) en la etapa 7 (Pago). En todas las etapas el conteo arranca **el día siguiente** a la fecha inicial (la fecha inicial cuenta como día 0).

**Avisos EXCLUIDOS (no entran al semáforo):** OBSERVACIÓN/DUPLICIDAD = REPETIDO, NULO (incl. OBS 03) o SIN COBERTURA (incl. OBS 08).

| # | Etapa | Verde | Ámbar | Rojo |
|---|-------|-------|-------|------|
| 1 | **Atención** | ≤6 días | 7-10 días | >10 días |
| 2 | **Programación** | ≤11 días | 12-15 días | >15 días |
| 3 | **Ajuste 01** | ≤11 días | 12-15 días | >15 días |
| 4 | **Reprogramación** | reprog. ya pasó / faltan ≤3 | faltan 4-7 días | faltan >7 días |
| 5 | **Padrón** | ≤15 días | 16-20 días | >20 días |
| 6 | **Validación** | ≤6 días | 7-15 días | >15 días |
| 7 | **Pago SAC** (días hábiles) | ≤11 días | 12-15 días | >15 días |

**Nota:** El motor calcula 14 columnas internas (`ALERTA_01..07` + `SEMAFORO_01..07`).
Etapas 1-6: port fiel de las fórmulas del Excel oficial (reconciliación fila-a-fila 100%).
Etapa 7 (Pago): desde julio 2026 cuenta los días hábiles **desde el día siguiente a la
validación** (decisión del equipo SAC; el Excel original contaba el día inicial y mostraba
un día de más).
        """)

    df = datos.get("midagri")
    if df is None or df.empty:
        st.warning("No hay datos cargados.")
        return

    if not check_semaforo_columns(df):
        st.error(
            "**Datos insuficientes para el Semáforo de Alertas**\n\n"
            "Los datos cargados no contienen las columnas necesarias para el control de plazos. "
            "Para habilitar esta funcionalidad, suba el archivo Excel expandido que incluye:\n"
            "- Fecha de Atención\n"
            "- Fecha de Programación de Ajuste\n"
            "- Fecha Ajuste Acta\n"
            "- Fechas de Reprogramación\n"
            "- Fecha Envío DRAS\n"
            "- Fecha Validación/Conformidad\n"
            "- Fecha Desembolso"
        )
        return

    # ── Fecha de corte (equivalente a la celda "FECHA DE REPORTE" del Excel) ──
    col_fc, col_info = st.columns([1, 2.4])
    with col_fc:
        corte_sel = st.date_input(
            "Fecha de corte",
            value=pd.Timestamp.now().date(),
            format="DD/MM/YYYY",
            help="Congela el cálculo a una fecha, igual que la celda "
                 "'FECHA DE REPORTE' del Excel. Para reproducir un R1 exacto, "
                 "ponla en la misma fecha de corte de ese reporte.",
            key="sem_fecha_corte")
    today = pd.Timestamp(corte_sel).normalize()
    with col_info:
        st.markdown(
            f'<div class="sem-corte-info"><span class="ms">event</span>'
            f'<div><b>{len(df):,} avisos</b> al corte '
            f'<b>{today.strftime("%d/%m/%Y")}</b><br>'
            f'<span>Días calendario en etapas 1-6 · días hábiles en etapa 7 '
            f'(Pago) · conteo desde el día siguiente a la fecha inicial</span></div></div>',
            unsafe_allow_html=True)

    # ── Cálculo ──
    _ck = (datos.get("fecha_corte", ""), datos.get("total_avisos", 0),
           today.strftime("%Y-%m-%d"))
    df_sem = compute_semaforo(df, today, cache_key=_ck)
    pipeline = get_pipeline_summary(df_sem)
    kpis = get_kpi_summary(df_sem)

    # ── KPIs ──
    st.markdown(_render_kpis_html(kpis), unsafe_allow_html=True)

    # ── Estado por etapa (conteo INDEPENDIENTE por etapa = hoja R1 del Excel) ──
    st.markdown("#### Estado por Etapa "
                "<span style='font-size:12px;font-weight:400;color:#7c8a82'>"
                "— conteo independiente, coincide con R1 del Excel</span>",
                unsafe_allow_html=True)
    st.markdown(_render_pipeline_html(pipeline), unsafe_allow_html=True)

    # ── Diagrama Sankey ──
    try:
        fig_sankey = generate_sankey_figure(df_sem)
        if fig_sankey:
            with st.expander("Diagrama de Flujo (Sankey)", expanded=False):
                try:
                    from shared.charts import render_chart
                    render_chart(fig_sankey, key="chart_sankey_semaforo",
                                 filename="semaforo_flujo_avisos")
                except ImportError:
                    st.plotly_chart(fig_sankey, use_container_width=True)
    except Exception:
        pass

    st.divider()

    # ── Filtros ──
    st.markdown("#### Filtros")
    c1, c2, c3, c4 = st.columns(4)

    deptos_list = sorted(df_sem["DEPARTAMENTO"].dropna().unique().tolist()) if "DEPARTAMENTO" in df_sem.columns else []
    empresas_list = sorted(df_sem["EMPRESA"].dropna().unique().tolist()) if "EMPRESA" in df_sem.columns else []
    etapas_opts = ["Todas"] + [s["label"] for s in STAGES]
    alertas_opts = ["verde", "ambar", "rojo"]

    with c1:
        fil_depto = st.multiselect("Departamento", deptos_list, key="sem_fil_depto")
    with c2:
        fil_empresa = st.multiselect("Empresa", empresas_list, key="sem_fil_empresa")
    with c3:
        fil_etapa = st.selectbox("Etapa", etapas_opts, key="sem_fil_etapa")
    with c4:
        fil_alerta = st.multiselect("Nivel de Alerta", alertas_opts,
                                     default=alertas_opts, key="sem_fil_alerta",
                                     format_func=lambda x: {"verde": "🟢 Verde", "ambar": "🟡 Ámbar", "rojo": "🔴 Rojo"}[x])

    # Aplicar filtros (sin .copy() innecesario — solo lectura)
    df_fil = df_sem[df_sem["SEM_ETAPA"] != "completado"]
    if fil_depto:
        df_fil = df_fil[df_fil["DEPARTAMENTO"].isin(fil_depto)]
    if fil_empresa:
        df_fil = df_fil[df_fil["EMPRESA"].isin(fil_empresa)]
    if fil_etapa != "Todas":
        stage_key = next((s["key"] for s in STAGES if s["label"] == fil_etapa), None)
        if stage_key:
            df_fil = df_fil[df_fil["SEM_ETAPA"] == stage_key]
    if fil_alerta:
        df_fil = df_fil[df_fil["SEM_ALERTA"].isin(fil_alerta)]

    st.markdown(f"**{len(df_fil):,} avisos** con alertas activas "
                f"(de {kpis['total']:,} totales)")

    # ── Tabla detalle ──
    st.markdown("#### Detalle de Alertas")

    display_cols = ["CODIGO_AVISO", "DEPARTAMENTO", "PROVINCIA", "DISTRITO",
                    "SECTOR_ESTADISTICO", "TIPO_CULTIVO", "EMPRESA",
                    "TIPO_SINIESTRO", "SEM_ETAPA", "SEM_ALERTA", "SEM_DIAS", "SEM_DETALLE"]
    available = [c for c in display_cols if c in df_fil.columns]

    df_display = df_fil[available].copy()
    df_display["SEM_ETAPA"] = df_display["SEM_ETAPA"].map(
        _STAGE_KEY_TO_LABEL).fillna(df_display["SEM_ETAPA"])

    rename_map = {
        "CODIGO_AVISO": "Código Aviso", "DEPARTAMENTO": "Departamento",
        "PROVINCIA": "Provincia", "DISTRITO": "Distrito",
        "SECTOR_ESTADISTICO": "Sector", "TIPO_CULTIVO": "Cultivo",
        "EMPRESA": "Empresa",
        "TIPO_SINIESTRO": "Tipo Siniestro", "SEM_ETAPA": "Etapa",
        "SEM_ALERTA": "Alerta", "SEM_DIAS": "Días", "SEM_DETALLE": "Detalle"
    }
    df_display = df_display.rename(columns={k: v for k, v in rename_map.items() if k in df_display.columns})

    st.dataframe(
        df_display,
        use_container_width=True,
        height=450,
        column_config={
            "Alerta": st.column_config.TextColumn(
                "Alerta", help="🟢 Verde: en plazo | 🟡 Ámbar: riesgo | 🔴 Rojo: vencido"),
            "Días": st.column_config.NumberColumn("Días", format="%d"),
        }
    )

    # ── Drill-down por etapa ──
    st.markdown("#### Análisis por Etapa")

    for s in STAGES:
        k = s["key"]
        p = pipeline.get(k, {"verde": 0, "ambar": 0, "rojo": 0, "total": 0})
        if p["total"] == 0:
            continue

        # Conteo INDEPENDIENTE de esta etapa (consistente con las tarjetas/R1)
        sem_col = STAGE_SEM_COL[k]
        alert_col = "ALERTA_" + sem_col.split("_")[1]
        vals = pd.to_numeric(df_sem[sem_col], errors="coerce")
        sub = df_sem[vals.isin([1, 2, 3])]      # avisos con alerta activa aquí

        with st.expander(f'{s["emoji"]} {s["label"]} — {p["total"]:,} alertas '
                         f'(🟢{p["verde"]:,} 🟡{p["ambar"]:,} 🔴{p["rojo"]:,})'):
            if "DEPARTAMENTO" in df_sem.columns:
                top_rojo = (df_sem[vals == 3]
                            .groupby("DEPARTAMENTO").size()
                            .sort_values(ascending=False).head(5))
                if not top_rojo.empty:
                    st.markdown('<div class="sem-drilldown-title"><span class="ms" style="color:var(--color-danger);">priority_high</span> Top 5 departamentos con más alertas rojas</div>',
                                unsafe_allow_html=True)
                    for depto, cnt in top_rojo.items():
                        st.markdown(f"&nbsp;&nbsp;&nbsp;**{depto}**: {cnt} avisos")

                top_total = (sub.groupby("DEPARTAMENTO").size()
                             .sort_values(ascending=False).head(5))
                if not top_total.empty:
                    st.markdown('<div class="sem-drilldown-title">Top 5 departamentos por volumen</div>',
                                unsafe_allow_html=True)
                    for depto, cnt in top_total.items():
                        st.markdown(f"&nbsp;&nbsp;&nbsp;**{depto}**: {cnt} avisos")

            # Promedio de días de ESTA etapa (extraído del texto "(N días)")
            dias = (df_sem.loc[sub.index, alert_col].astype(str)
                    .str.extract(r"\((-?\d+)\s*d", expand=False))
            dias = pd.to_numeric(dias, errors="coerce")
            if dias.notna().any():
                st.markdown(f"**Promedio de días en esta etapa:** {dias.mean():.1f} días")

    # ── Exportar ──
    st.divider()
    col_exp1, col_exp2 = st.columns([1, 3])
    with col_exp1:
        if st.button("Generar Excel Semáforo", key="sem_gen_excel", type="primary"):
            excel_bytes = export_semaforo_excel(df_fil, pipeline, kpis)
            st.session_state["sem_excel"] = excel_bytes
            st.session_state["sem_excel_name"] = f"semaforo_alertas_{today.strftime('%d%m%Y')}.xlsx"

    with col_exp2:
        if st.session_state.get("sem_excel"):
            st.download_button(
                ":material/download: Descargar Excel Semáforo",
                data=st.session_state["sem_excel"],
                file_name=st.session_state["sem_excel_name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="sem_download_excel"
            )
